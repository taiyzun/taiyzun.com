from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook


FARHAT = Path("/Users/tai/Downloads/BFP Main List - May 18 3am Share File Farhat.xlsx")
OUT = Path("/Users/tai/Documents/GitHub/taiyzun.com/outputs/bfp_farhat_may18_review")
OUT.mkdir(parents=True, exist_ok=True)

HONORIFICS = [
    "Hon'ble Shri",
    "Hon'ble Smt",
    "Hon'ble Mr",
    "Hon'ble",
    "Most Venerable Dr",
    "Most Venerable",
    "Respected Sardar",
    "Professor",
    "Prof.",
    "Prof",
    "Padma Shri",
    "H.E Dr",
    "H.E.",
    "H.E",
    "HE",
    "Dr.",
    "Dr",
    "Mr.",
    "Mr",
    "Mrs.",
    "Mrs",
    "Ms.",
    "Ms",
    "Adv.",
    "Adv",
    "Shri",
    "Smt",
]
HONORIFICS = sorted(HONORIFICS, key=len, reverse=True)


def clean(value) -> str:
    return "" if value is None else str(value).replace("\u00a0", " ").strip()


def normalise_name(value: str) -> str:
    text = clean(value).lower()
    text = re.sub(r"^[\s•⁠\-–—]+", "", text)
    text = re.sub(r"\([^)]*\)", " ", text)
    text = text.replace("&", " and ")
    text = re.sub(
        r"^(hon'?ble shri|hon'?ble smt|hon'?ble mr|hon'?ble|most venerable dr|most venerable|respected sardar|professor|prof|padma shri|h\.e dr|h\.e\.|h\.e|he|dr|mr|mrs|ms|adv|shri|smt)\.?\s+",
        "",
        text,
    )
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def split_name(value: str) -> tuple[str, str]:
    text = clean(value).lstrip("•⁠-–— ").strip()
    for honorific in HONORIFICS:
        if text.lower().startswith(honorific.lower() + " "):
            return honorific, text[len(honorific) :].strip()
    return "", text


def phone(value: str) -> str:
    digits = re.sub(r"\D+", "", clean(value))
    if not digits:
        return ""
    if len(digits) == 10:
        return "+91" + digits
    if len(digits) < 10:
        return ""
    return "+" + digits


def seat(value: str) -> str:
    raw = clean(value)
    if re.fullmatch(r"sofa\s*-?", raw, re.I):
        return "Sofa"
    if re.fullmatch(r"\d+\s*-", raw):
        return raw.replace(" ", "").replace("-", "")
    return raw


def status_for(row: dict) -> tuple[str, str]:
    category_final = row.get("Category Final", "")
    category = row.get("Category", "")
    award = row.get("Award Category (where applicable)", "")
    if category_final == "Guest of Honour":
        if category == "Awardee" or award:
            return "I Am Peacekeeper Champion", award
        return "Guest of Honour", ""
    if category_final == "VIP Peace Partner":
        return "VIP I Am Peacekeeper Partner", ""
    if category_final == "Main":
        if category == "Chief Guest":
            return "Chief Guest", ""
        if category == "Nobel Laureate":
            return "Nobel Peace Laureates", ""
        return "KEEP_EXISTING_OR_MANUAL", ""
    if category_final == "I Am Peacekeeper":
        return "KEEP_EXISTING_IAP_STATUS", ""
    if category_final == "Late Peace Guest":
        return "Late Peace Guest/manual", ""
    if category_final == "Media":
        return "Media/manual only", ""
    if category_final == "XXX":
        return "DO_NOT_UPDATE_OR_ADD", ""
    return "MANUAL", ""


def skip_reason(row: dict, display_name: str, target_status: str) -> str:
    category_final = row.get("Category Final", "")
    lower = display_name.lower()
    if category_final == "XXX":
        return "XXX row"
    if category_final == "Media":
        return "Media row"
    if re.fullmatch(r"(gail|petronet)\s+person", lower):
        return "placeholder/no real guest name"
    if "xxx" in lower or lower in {"media", "person"}:
        return "placeholder/no real guest name"
    if target_status == "Late Peace Guest/manual":
        return "Late Peace Guest tier is not available in system dropdown"
    return ""


wb = load_workbook(FARHAT, data_only=True, read_only=True)
ws = wb.active
headers = [cell.value if cell.value is not None else f"__blank_{i + 1}" for i, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))]

targets = []
for excel_row, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    row = {str(headers[i]): clean(values[i] if i < len(values) else "") for i in range(len(headers))}
    display_name = clean(row.get("Name"))
    if not display_name:
        continue
    honorific, name = split_name(display_name)
    target_status, champion_category = status_for(row)
    skip = skip_reason(row, display_name, target_status)
    sort_name = clean(row.get("Name For Sort"))
    search_queries = [display_name, name]
    if sort_name:
        search_queries.append(sort_name)
    parts = name.split()
    if len(parts) >= 2:
        search_queries.append(" ".join(parts[-2:]))
    if parts:
        search_queries.append(parts[-1])
    search_queries = list(dict.fromkeys(q for q in search_queries if q))
    targets.append(
        {
            "fans_row": str(excel_row),
            "fans_name": display_name,
            "honorific": honorific,
            "name": name,
            "normalised_name": normalise_name(name or display_name),
            "category_final": row.get("Category Final", ""),
            "category": row.get("Category", ""),
            "award_category_farhat": row.get("Award Category (where applicable)", ""),
            "target_status": target_status,
            "target_champion_category": champion_category if target_status == "I Am Peacekeeper Champion" else "",
            "seat_18th_3am": seat(row.get("Seat 18th 3am", "")),
            "phone": phone(row.get("Cell", "")),
            "email": row.get("email", ""),
            "designation_company": row.get("Designation, Company", ""),
            "country": row.get("India / International", ""),
            "skip_reason": skip,
            "search_queries": search_queries,
        }
    )

active = [row for row in targets if not row["skip_reason"]]
payload = {
    "source": str(FARHAT),
    "total_named_rows": len(targets),
    "active_rows_to_live_check": len(active),
    "skipped_rows": len(targets) - len(active),
    "targets": targets,
}

(OUT / "farhat_full_missing_live_targets.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
print(json.dumps({k: payload[k] for k in ("total_named_rows", "active_rows_to_live_check", "skipped_rows")}, indent=2))
