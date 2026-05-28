from openpyxl import load_workbook
from pathlib import Path
from collections import Counter, defaultdict
import csv, re, json, difflib

NEW = Path('/Users/tai/Downloads/BFP Main List - May 18 3am Share File Farhat.xlsx')
OLD = Path('/Users/tai/Library/Mobile Documents/com~apple~CloudDocs/Work/Chin/Chin/BFP Main FANS List - WIP May 17 2pm File.xlsx')
SYS = Path('/Users/tai/Downloads/guests-2026-05-17-658.csv')
OUT = Path('/Users/tai/Documents/GitHub/taiyzun.com/outputs/bfp_farhat_may18_review')
OUT.mkdir(parents=True, exist_ok=True)

HON = re.compile(r"^(mr|mrs|ms|miss|dr|prof|professor|shri|smt|hon'?ble|hon|he|h\.e\.|h\.h\.|hh|sr\.?|fr\.?|rev\.?|ambassador|maharaja|princess|prince)\.?\s+", re.I)
BUL = re.compile(r"^[\s•⁠\-–—]+")

def clean(x):
    if x is None: return ''
    return str(x).replace('\u00a0',' ').strip()

def norm_name(x):
    s=clean(x).lower()
    s=BUL.sub('',s)
    s=s.replace('&',' and ')
    s=re.sub(r"\([^)]*\)"," ",s)
    for _ in range(4): s=HON.sub('',s).strip()
    s=re.sub(r"[^a-z0-9]+"," ",s)
    s=re.sub(r"\s+"," ",s).strip()
    return s

def row_key(row):
    return norm_name(row.get('Name For Sort') or row.get('Name'))

def load_fans(path):
    wb=load_workbook(path, data_only=True, read_only=True)
    ws=wb[wb.sheetnames[0]]
    headers=[c.value if c.value is not None else f'__blank_{i+1}' for i,c in enumerate(next(ws.iter_rows(min_row=1,max_row=1)))]
    rows=[]
    for idx,r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        d={headers[i]: clean(r[i]) if i < len(r) else '' for i in range(len(headers))}
        if any(d.values()):
            d['_row']=idx; d['_key']=row_key(d); d['_display']=clean(d.get('Name'))
            rows.append(d)
    return rows, headers, ws.title

def load_system(path):
    with open(path, newline='', encoding='utf-8-sig') as f:
        rows=list(csv.DictReader(f))
    for d in rows:
        d['_key']=norm_name(d.get('name'))
    return rows

new_rows,new_headers,new_sheet=load_fans(NEW)
old_rows,old_headers,old_sheet=load_fans(OLD)
sys_rows=load_system(SYS)

def by_key(rows):
    m=defaultdict(list)
    for r in rows:
        if r['_key']: m[r['_key']].append(r)
    return m
new_by=by_key(new_rows); old_by=by_key(old_rows); sys_by=by_key(sys_rows)

# counts
summary=[]
def add_summary(metric, value, note=''):
    summary.append({'Metric': metric, 'Value': value, 'Note': note})
add_summary('New workbook rows with names', sum(1 for r in new_rows if r.get('Name')), 'All named rows, including XXX and Media')
add_summary('Rows marked XXX', sum(1 for r in new_rows if r.get('Category Final')=='XXX'), '')
add_summary('Non-XXX named rows', sum(1 for r in new_rows if r.get('Name') and r.get('Category Final')!='XXX'), 'Farhat said 505 total; file currently gives 504 if XXX rows are excluded')
add_summary('VIP Peace Partner rows', sum(1 for r in new_rows if r.get('Category Final')=='VIP Peace Partner'), '')
add_summary('Guest of Honour rows', sum(1 for r in new_rows if r.get('Category Final')=='Guest of Honour'), '')
add_summary('I Am Peacekeeper rows', sum(1 for r in new_rows if r.get('Category Final')=='I Am Peacekeeper'), '')
add_summary('Media rows', sum(1 for r in new_rows if r.get('Category Final')=='Media'), '')
add_summary('Awardee rows', sum(1 for r in new_rows if r.get('Category')=='Awardee'), 'Farhat said 180 awards; file currently has 186 Awardee rows')
add_summary('Rows with award category filled', sum(1 for r in new_rows if r.get('Award Category (where applicable)')), 'Includes Business XXX / Justice & Peace XXX / typo values')
add_summary('Award category filled and not XXX final', sum(1 for r in new_rows if r.get('Award Category (where applicable)') and r.get('Category Final')!='XXX'), '')
add_summary('Awardee rows with no award category', sum(1 for r in new_rows if r.get('Category')=='Awardee' and not r.get('Award Category (where applicable)')), '')
add_summary('IAP rows with a 3am seat value', sum(1 for r in new_rows if r.get('Category Final')=='I Am Peacekeeper' and r.get('Seat 18th 3am')), 'Contradicts verbal “no place to sit for all IAPMs” only if seat values are meant to be blank')

# added/removed unique keys
new_keys=set(new_by); old_keys=set(old_by)
added_keys=sorted(new_keys-old_keys)
removed_keys=sorted(old_keys-new_keys)
added=[new_by[k][0] for k in added_keys]
removed=[old_by[k][0] for k in removed_keys]

# field changes for one-to-one common keys
fields=['Name','Designation, Company','Category Final','Seat 18th 3am','Seat 17th 2pm','Category','Award Category (where applicable)','Speaking','Connected by','India / International','Cell','email','Confirmed after 12th May']
old_field_map={'Seat 18th 3am':'Seat 2pm','Seat 17th 2pm':'Seat 2pm','Speaking':None,'India / International':None}
changes=[]
for k in sorted(new_keys & old_keys):
    if len(new_by[k])!=1 or len(old_by[k])!=1:
        continue
    n=new_by[k][0]; o=old_by[k][0]
    diffs=[]
    for f in fields:
        of=old_field_map.get(f,f)
        if of is None:
            ov=''
        else:
            ov=clean(o.get(of,''))
        nv=clean(n.get(f,''))
        # do not flag new Seat 17th if just carrying old Seat 2pm identical unless Seat18 changed tracked separately
        if f=='Seat 17th 2pm':
            continue
        if nv!=ov:
            diffs.append((f,ov,nv))
    if diffs:
        for f,ov,nv in diffs:
            changes.append({'Key':k,'Name':n.get('Name'),'Row May18':n['_row'],'Field':f,'May17':ov,'May18':nv})

# category/status change subset
cat_changes=[c for c in changes if c['Field'] in ('Category Final','Category','Award Category (where applicable)','Seat 18th 3am','Cell','email','Designation, Company','Speaking')]

# issues/fuzzy
issues=[]
controlled={'Artist & Literature','Business','City','Education','Environment','Health','Housing','Influencers','Justice & Peace','Social Innovation','Sports','Youth & Women'}
for r in new_rows:
    name=r.get('Name')
    cf=r.get('Category Final')
    cat=r.get('Category')
    award=r.get('Award Category (where applicable)')
    seat=r.get('Seat 18th 3am')
    if cf=='I Am Peacekeeper' and seat:
        issues.append({'Issue':'IAP has seat value','Name':name,'Row':r['_row'],'Category Final':cf,'Category':cat,'Award Category':award,'Seat 18th 3am':seat,'Suggested action':'Confirm whether to blank seat in system; do not add seat.'})
    if cat=='Awardee' and not award:
        issues.append({'Issue':'Awardee missing award category','Name':name,'Row':r['_row'],'Category Final':cf,'Category':cat,'Award Category':award,'Seat 18th 3am':seat,'Suggested action':'Ask Farhat / user for Champion Category before update.'})
    if award and award not in controlled:
        suggestion = {'Champion Business':'Business','Peace':'Justice & Peace'}.get(award,'')
        if 'XXX' in award:
            suggestion='Manual review: category contains XXX marker'
        issues.append({'Issue':'Award category wording is not controlled','Name':name,'Row':r['_row'],'Category Final':cf,'Category':cat,'Award Category':award,'Seat 18th 3am':seat,'Suggested action':suggestion or 'Confirm corrected wording before update.'})
    if cf in ('XXX','Media'):
        issues.append({'Issue':f'{cf} row should not be auto-added/updated','Name':name,'Row':r['_row'],'Category Final':cf,'Category':cat,'Award Category':award,'Seat 18th 3am':seat,'Suggested action':'Manual review only.'})
    if re.search(r'who\?|check|xxx|unclear|tbc|to be confirmed', name+' '+cat+' '+cf, re.I):
        issues.append({'Issue':'Unclear marker in row','Name':name,'Row':r['_row'],'Category Final':cf,'Category':cat,'Award Category':award,'Seat 18th 3am':seat,'Suggested action':'Manual review.'})

# duplicates in new
for k,rs in sorted(new_by.items()):
    if len(rs)>1:
        issues.append({'Issue':'Duplicate/same-normalised-name in Farhat list','Name':' | '.join(r.get('Name') for r in rs),'Row':'; '.join(str(r['_row']) for r in rs),'Category Final':' | '.join(r.get('Category Final') for r in rs),'Category':' | '.join(r.get('Category') for r in rs),'Award Category':' | '.join(r.get('Award Category (where applicable)') for r in rs),'Seat 18th 3am':' | '.join(r.get('Seat 18th 3am') for r in rs),'Suggested action':'Confirm if duplicate people or separate guests before system update.'})

# compare against system csv for missing/fuzzy by name
fans_active=[r for r in new_rows if r.get('Name') and r.get('Category Final') not in ('XXX','Media')]
missing=[]; fuzzy=[]; exact=0
sys_keys=set(sys_by)
for r in fans_active:
    k=r['_key']
    if k in sys_keys:
        exact += 1; continue
    # best fuzzy
    candidates=difflib.get_close_matches(k, list(sys_keys), n=3, cutoff=0.84)
    if candidates:
        fuzzy.append({'FANS Name':r.get('Name'),'FANS Row':r['_row'],'Category Final':r.get('Category Final'),'Category':r.get('Category'),'Best System Matches':' | '.join(sys_by[c][0].get('name','') for c in candidates),'Match Keys':' | '.join(candidates)})
    else:
        missing.append({'FANS Name':r.get('Name'),'FANS Row':r['_row'],'Category Final':r.get('Category Final'),'Category':r.get('Category'),'Award Category':r.get('Award Category (where applicable)'),'Cell':r.get('Cell'),'email':r.get('email')})

# Moonmoon check
moon=[r for r in new_rows if re.search(r'moon\s*moon|moonmoon|sen', r.get('Name',''), re.I)]
moon_sys=[]
for r in sys_rows:
    if re.search(r'moon\s*moon|moonmoon|sen', r.get('name',''), re.I): moon_sys.append(r)

# write csvs
def write_csv(name, rows, headers=None):
    p=OUT/name
    if not rows:
        rows=[]
        if headers is None: headers=['Note']
    if headers is None:
        headers=[]
        for row in rows:
            for k in row:
                if k not in headers: headers.append(k)
    with open(p,'w',newline='',encoding='utf-8-sig') as f:
        w=csv.DictWriter(f, fieldnames=headers, extrasaction='ignore')
        w.writeheader(); w.writerows(rows)

write_csv('summary.csv', summary)
write_csv('added_in_farhat_may18.csv', added)
write_csv('removed_since_may17.csv', removed)
write_csv('field_changes_may17_to_may18.csv', changes)
write_csv('important_changes_may17_to_may18.csv', cat_changes)
write_csv('fuzzy_manual_review_farhat_may18.csv', issues)
write_csv('fans_not_exact_in_system_csv.csv', missing)
write_csv('fans_fuzzy_matches_system_csv.csv', fuzzy)
write_csv('moonmoon_sen_check_fans.csv', moon)
write_csv('moonmoon_sen_check_system_csv.csv', moon_sys)

# JSON compact for final
payload={
 'summary': summary,
 'category_final_counts': Counter(r.get('Category Final') for r in new_rows if r.get('Category Final')).most_common(),
 'category_counts': Counter(r.get('Category') for r in new_rows if r.get('Category')).most_common(),
 'award_counts': Counter(r.get('Award Category (where applicable)') for r in new_rows if r.get('Award Category (where applicable)')).most_common(),
 'added_count': len(added), 'removed_count': len(removed), 'changes_count': len(changes), 'important_changes_count': len(cat_changes),
 'issues_count': len(issues), 'missing_in_system_csv_count': len(missing), 'fuzzy_system_count': len(fuzzy), 'system_exact_name_count': exact,
 'added_sample': [{k:r.get(k,'') for k in ['Name','Category Final','Category','Award Category (where applicable)','Seat 18th 3am','Cell','email']} for r in added[:30]],
 'removed_sample': [{k:r.get(k,'') for k in ['Name','Category Final','Category','Award Category (where applicable)','Seat 2pm','Cell','email']} for r in removed[:30]],
 'important_changes_sample': cat_changes[:60],
 'issues_sample': issues[:80],
 'moon_fans': [{k:r.get(k,'') for k in ['Name','Category Final','Category','Award Category (where applicable)','Seat 18th 3am','Cell','email']} for r in moon],
 'moon_system': [{k:r.get(k,'') for k in ['id','name','status_label','phone','email','champion_category','award_category']} for r in moon_sys],
}
(OUT/'analysis_summary.json').write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')
print(json.dumps(payload, ensure_ascii=False, indent=2)[:12000])
