from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


BASE = Path("/Users/tai/Documents/GitHub/taiyzun.com/outputs/bfp_farhat_may18_review")
FARHAT = Path("/Users/tai/Downloads/BFP Main List - May 18 3am Share File Farhat.xlsx")
SYSTEM = Path("/Users/tai/Downloads/guests-2026-05-17-658.csv")
LIVE = BASE / "farhat_full_missing_live_check_corrected.json"

HONORIFIC_RE = re.compile(
    r"^(hon'?ble shri|hon'?ble smt|hon'?ble mr|hon'?ble|most venerable dr|most venerable|respected sardar|professor|prof|padma shri|h\.e dr|h\.e\.|h\.e|he|dr|mr|mrs|ms|adv|shri|smt)\.?\s+",
    re.I,
)


def clean(value) -> str:
    return "" if value is None else str(value).replace("\u00a0", " ").strip()


def norm_name(value: str) -> str:
    text = clean(value).lower()
    text = re.sub(r"^[\s•⁠\-–—]+", "", text)
    text = re.sub(r"\([^)]*\)", " ", text)
    text = text.replace("&", " and ")
    for _ in range(4):
        text = HONORIFIC_RE.sub("", text).strip()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def norm_phone(value: str) -> str:
    digits = re.sub(r"\D+", "", clean(value))
    if not digits:
        return ""
    if len(digits) == 10:
        return "+91" + digits
    return "+" + digits


def norm_email(value: str) -> str:
    return clean(value).lower()


def is_placeholder_name(name: str, category_final: str) -> bool:
    lower = clean(name).lower()
    return (
        category_final in {"XXX", "Media"}
        or re.fullmatch(r"(gail|petronet)\s+person", lower) is not None
        or "xxx" in lower
        or lower in {"media", "person"}
    )


def load_farhat() -> list[dict]:
    wb = load_workbook(FARHAT, data_only=True, read_only=True)
    ws = wb.active
    headers = [cell.value if cell.value is not None else f"__blank_{idx + 1}" for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))]
    rows = []
    for excel_row, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row = {str(headers[i]): clean(values[i] if i < len(values) else "") for i in range(len(headers))}
        if not row.get("Name"):
            continue
        row["_row"] = str(excel_row)
        row["_norm_name"] = norm_name(row.get("Name For Sort") or row.get("Name"))
        row["_norm_phone"] = norm_phone(row.get("Cell"))
        row["_norm_email"] = norm_email(row.get("email"))
        row["_skip_placeholder"] = is_placeholder_name(row.get("Name", ""), row.get("Category Final", ""))
        rows.append(row)
    return rows


def load_system() -> list[dict]:
    with SYSTEM.open(newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.DictReader(handle))
    for row in rows:
        row["_norm_name"] = norm_name(row.get("name", ""))
        row["_norm_phone"] = norm_phone(row.get("phone", ""))
        row["_norm_email"] = norm_email(row.get("email", ""))
    return rows


def grouped_duplicates(rows: list[dict], key: str) -> dict[str, list[dict]]:
    groups: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        value = row.get(key, "")
        if value:
            groups[value].append(row)
    return {value: items for value, items in groups.items() if len(items) > 1}


def write_csv(path: Path, rows: list[dict], headers: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def farhat_dup_rows(farhat: list[dict]) -> list[dict]:
    active = [row for row in farhat if not row["_skip_placeholder"]]
    output = []
    for duplicate_type, key in [("name", "_norm_name"), ("phone", "_norm_phone"), ("email", "_norm_email")]:
        for value, items in grouped_duplicates(active, key).items():
            output.append(
                {
                    "duplicate_type": duplicate_type,
                    "normalised_value": value,
                    "count": len(items),
                    "fans_rows": " | ".join(item["_row"] for item in items),
                    "fans_names": " | ".join(item.get("Name", "") for item in items),
                    "category_final": " | ".join(item.get("Category Final", "") for item in items),
                    "category": " | ".join(item.get("Category", "") for item in items),
                    "phones": " | ".join(item.get("Cell", "") for item in items),
                    "emails": " | ".join(item.get("email", "") for item in items),
                    "recommendation": "Manual review before add/update; do not create separate records unless these are confirmed separate people.",
                }
            )
    return output


def system_dup_rows(system: list[dict]) -> list[dict]:
    output = []
    for duplicate_type, key in [("system_id", "id"), ("name", "_norm_name"), ("phone", "_norm_phone"), ("email", "_norm_email")]:
        for value, items in grouped_duplicates(system, key).items():
            # Ignore blank-like phone/email already filtered by grouped_duplicates, but keep placeholders: they matter.
            output.append(
                {
                    "duplicate_type": duplicate_type,
                    "normalised_value": value,
                    "count": len(items),
                    "system_ids": " | ".join(item.get("id", "") for item in items),
                    "system_names": " | ".join(item.get("name", "") for item in items),
                    "statuses": " | ".join(item.get("status_label", "") or item.get("status", "") for item in items),
                    "phones": " | ".join(item.get("phone", "") for item in items),
                    "emails": " | ".join(item.get("email", "") for item in items),
                    "recommendation": "Live confirmation required before merge/delete. Never delete system-only records automatically.",
                }
            )
    return output


def live_signal_rows() -> tuple[list[dict], list[dict], list[dict]]:
    if not LIVE.exists():
        return [], [], []
    data = json.loads(LIVE.read_text(encoding="utf-8"))
    duplicate_live = []
    fuzzy = []
    phone_conflicts = []
    for row in data["results"]:
        matches = row.get("primary_matches", [])
        match_text = " | ".join(f"{m.get('text','')} <{m.get('id','')}>" for m in matches)
        common = {
            "fans_row": row.get("fans_row", ""),
            "fans_name": row.get("fans_name", ""),
            "target_status": row.get("target_status", ""),
            "category_final": row.get("category_final", ""),
            "category": row.get("category", ""),
            "seat_18th_3am": row.get("seat_18th_3am", ""),
            "phone": row.get("phone", ""),
            "matches": match_text,
        }
        status = row.get("status")
        if status in {"duplicate_exact_live_matches", "duplicate_phone_live_matches"}:
            duplicate_live.append({**common, "status": status, "recommendation": "Confirmed live duplicate signal; show to user before merge/delete."})
        elif status == "fuzzy_or_related_matches":
            fuzzy.append({**common, "status": status, "recommendation": "Fuzzy/related only. Do not add or merge without confirmation."})
        elif status == "phone_live_match":
            phone_conflicts.append({**common, "status": status, "recommendation": "Phone matched a different or variant live name. Verify before update."})
    return duplicate_live, fuzzy, phone_conflicts


farhat = load_farhat()
system = load_system()
farhat_dups = farhat_dup_rows(farhat)
system_dups = system_dup_rows(system)
live_dups, live_fuzzy, live_phone = live_signal_rows()

dup_headers = [
    "duplicate_type",
    "normalised_value",
    "count",
    "fans_rows",
    "fans_names",
    "category_final",
    "category",
    "phones",
    "emails",
    "recommendation",
]
sys_headers = [
    "duplicate_type",
    "normalised_value",
    "count",
    "system_ids",
    "system_names",
    "statuses",
    "phones",
    "emails",
    "recommendation",
]
live_headers = ["fans_row", "fans_name", "status", "target_status", "category_final", "category", "seat_18th_3am", "phone", "matches", "recommendation"]

write_csv(BASE / "duplicate_checks_farhat_duplicates_2026-05-18.csv", farhat_dups, dup_headers)
write_csv(BASE / "duplicate_checks_system_csv_duplicates_2026-05-18.csv", system_dups, sys_headers)
write_csv(BASE / "duplicate_checks_live_duplicate_signals_2026-05-18.csv", live_dups, live_headers)
write_csv(BASE / "duplicate_checks_live_fuzzy_review_2026-05-18.csv", live_fuzzy, live_headers)
write_csv(BASE / "duplicate_checks_live_phone_conflicts_2026-05-18.csv", live_phone, live_headers)

summary = {
    "farhat_rows": len(farhat),
    "farhat_active_rows": sum(1 for row in farhat if not row["_skip_placeholder"]),
    "farhat_duplicate_groups": len(farhat_dups),
    "farhat_duplicate_groups_by_type": dict(Counter(row["duplicate_type"] for row in farhat_dups)),
    "system_csv_rows": len(system),
    "system_csv_duplicate_groups": len(system_dups),
    "system_csv_duplicate_groups_by_type": dict(Counter(row["duplicate_type"] for row in system_dups)),
    "live_duplicate_signal_rows": len(live_dups),
    "live_fuzzy_review_rows": len(live_fuzzy),
    "live_phone_conflict_rows": len(live_phone),
    "high_priority_notes": [
        "Use live duplicate signals and fuzzy review before making any merge/delete decision.",
        "System CSV is an older snapshot; use it to find risk, then confirm live before edits.",
        "Farhat duplicate rows may intentionally represent one person appearing in multiple categories; confirm before deletion.",
    ],
}
(BASE / "duplicate_checks_summary_2026-05-18.json").write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
print(json.dumps(summary, indent=2, ensure_ascii=False))
