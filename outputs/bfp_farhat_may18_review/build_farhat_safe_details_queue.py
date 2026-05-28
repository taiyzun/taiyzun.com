import csv
import json
import re
from pathlib import Path

from openpyxl import load_workbook

BASE = Path("/Users/tai/Documents/GitHub/taiyzun.com/outputs/bfp_farhat_may18_review")
SRC = BASE / "farhat_may18_decision_queue.csv"
OUT = BASE / "farhat_may18_safe_details_queue.json"
HELD = BASE / "farhat_may18_safe_details_held.json"
FARHAT = Path("/Users/tai/Downloads/BFP Main List - May 18 3am Share File Farhat.xlsx")

SKIP_CATEGORY_FINAL = {"XXX", "Media"}
SAFE_MATCH_TYPES = {"exact_name", "phone"}


def clean(value: str | None) -> str:
    return "" if value is None else str(value).replace("\u00a0", " ").strip()


def phone_e164ish(value: str) -> str:
    raw = clean(value)
    if not raw:
        return ""
    digits = re.sub(r"\D+", "", raw)
    if not digits:
        return ""
    if len(digits) == 10:
        return "+91" + digits
    if raw.lstrip().startswith("+"):
        return "+" + digits
    return "+" + digits


def split_designation_company(value: str) -> tuple[str, str]:
    text = clean(value)
    if not text:
        return "", ""
    # Keep "with ..." notes as designation context, not organisation.
    if text.lower().startswith("with "):
        return text, ""
    parts = [p.strip() for p in text.split(",") if p.strip()]
    if len(parts) >= 2:
        return ", ".join(parts[:-1]), parts[-1]
    return text, ""


queue: list[dict[str, object]] = []
held: list[dict[str, object]] = []

wb = load_workbook(FARHAT, read_only=True, data_only=True)
ws = wb.active
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
farhat_by_row: dict[str, dict[str, str]] = {}
for row_no, values in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
    farhat_by_row[str(row_no)] = {
        str(headers[idx]): clean(values[idx] if idx < len(values) else "")
        for idx in range(len(headers))
        if headers[idx] is not None
    }

with SRC.open(encoding="utf-8-sig", newline="") as handle:
    for row in csv.DictReader(handle):
        system_id = clean(row["System ID"])
        match_type = clean(row["Match Type"])
        cf = clean(row["FANS Category Final"])
        cell = phone_e164ish(row["FANS Cell"])
        email = clean(row["FANS Email"])
        frow = farhat_by_row.get(row["FANS Row"], {})
        designation, organization = split_designation_company(frow.get("Designation, Company", ""))
        country = clean(frow.get("India / International", ""))
        if organization and country and organization.lower() == country.lower():
            organization = ""

        # The decision queue does not include the designation column, so this
        # pass reads it from the analysis detail files by row only when present.
        item = {
            "id": system_id,
            "fans_row": row["FANS Row"],
            "fans_name": row["FANS Name"],
            "match_type": match_type,
            "target": {
                "phone": cell,
                "email": email,
                "designation": designation,
                "organization": organization,
                "country": country,
            },
            "source": {
                "FANS Category Final": cf,
                "FANS Category": row["FANS Category"],
                "FANS Award Category": row["FANS Award Category"],
                "System Name": row["System Name"],
                "System Status": row["System Status"],
                "Reasons": row["Reasons"],
            },
        }
        if cf in SKIP_CATEGORY_FINAL:
            held.append({**item, "hold_reason": "XXX/media row"})
            continue
        if not system_id or "|" in system_id:
            held.append({**item, "hold_reason": "missing or multiple live System IDs"})
            continue
        if match_type not in SAFE_MATCH_TYPES:
            held.append({**item, "hold_reason": "not exact/phone match"})
            continue
        if not any([cell, email, designation, organization, country]):
            continue
        queue.append(item)

OUT.write_text(json.dumps(queue, ensure_ascii=False, indent=2), encoding="utf-8")
HELD.write_text(json.dumps(held, ensure_ascii=False, indent=2), encoding="utf-8")
print(json.dumps({"queue": len(queue), "held": len(held), "out": str(OUT), "sample": queue[:8]}, ensure_ascii=False, indent=2))
