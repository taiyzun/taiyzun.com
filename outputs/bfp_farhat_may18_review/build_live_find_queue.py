import csv
import json
import re
from pathlib import Path

from openpyxl import load_workbook

BASE = Path("/Users/tai/Documents/GitHub/taiyzun.com/outputs/bfp_farhat_may18_review")
SRC = BASE / "queue_ask_add_or_find_live.csv"
FARHAT = Path("/Users/tai/Downloads/BFP Main List - May 18 3am Share File Farhat.xlsx")
OUT = BASE / "farhat_may18_live_find_queue.json"

HON = [
    "Hon'ble Shri",
    "Hon'ble Smt",
    "Hon'ble",
    "Professor",
    "Prof.",
    "Prof",
    "H.E. Hon.",
    "H.E. Mr",
    "H.E.",
    "HE",
    "Dr.",
    "Dr",
    "Mr.",
    "Mr",
    "Mrs.",
    "Mrs",
    "Ms.",
    "Ms",
    "Shri",
    "Smt",
]
HON = sorted(HON, key=len, reverse=True)


def clean(v):
    return "" if v is None else str(v).replace("\u00a0", " ").strip()


def split_name(v):
    text = clean(v).lstrip("•⁠-–— ").strip()
    for h in HON:
        if text.lower().startswith(h.lower() + " "):
            return h, text[len(h) :].strip()
    return "", text


def phone(v):
    raw = clean(v)
    digits = re.sub(r"\D+", "", raw)
    if not digits:
        return ""
    if len(digits) == 10:
        return "+91" + digits
    if len(digits) < 10:
        return ""
    return "+" + digits


def seat(v):
    s = clean(v)
    if re.fullmatch(r"sofa\s*-?", s, re.I):
        return "Sofa"
    if re.fullmatch(r"(\d+)\s*-", s):
        return re.sub(r"\s*-", "", s)
    return s


def status_code(label):
    return {
        "Chief Guest": "chief_guest",
        "Nobel Peace Laureates": "nobel_laureate",
        "Guest of Honour": "guest_of_honour",
        "I Am Peacekeeper Champion": "champion",
        "VIP I Am Peacekeeper Partner": "vip_partner",
    }.get(label, "")


wb = load_workbook(FARHAT, read_only=True, data_only=True)
ws = wb.active
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
farhat_by_row = {}
for row_no, values in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
    farhat_by_row[str(row_no)] = {str(headers[i]): clean(values[i] if i < len(values) else "") for i in range(len(headers)) if headers[i] is not None}

queue = []
with SRC.open(encoding="utf-8-sig", newline="") as handle:
    for row in csv.DictReader(handle):
        cf = row["FANS Category Final"]
        if cf in {"XXX", "Media"}:
            continue
        honorific, name = split_name(row["FANS Name"])
        frow = farhat_by_row.get(row["FANS Row"], {})
        target = {
            "honorific": honorific,
            "name": name,
            "status_label": row["Target Status"],
            "status_code": status_code(row["Target Status"]),
            "seat_no": seat(row["FANS Seat 18th 3am"]),
            "award_category": "",
            "champion_category": row["Target Champion Category"] if row["Target Status"] == "I Am Peacekeeper Champion" else "",
            "phone": phone(row["FANS Cell"]),
            "email": row["FANS Email"],
            "designation_company": frow.get("Designation, Company", ""),
            "country": frow.get("India / International", ""),
        }
        queue.append(
            {
                "fans_row": row["FANS Row"],
                "fans_name": row["FANS Name"],
                "search_queries": [row["FANS Name"], name],
                "target": target,
                "source": {
                    "FANS Category Final": row["FANS Category Final"],
                    "FANS Category": row["FANS Category"],
                    "FANS Award Category": row["FANS Award Category"],
                    "Reasons": row["Reasons"],
                },
            }
        )

OUT.write_text(json.dumps(queue, ensure_ascii=False, indent=2), encoding="utf-8")
print(json.dumps({"queue": len(queue), "out": str(OUT), "sample": queue[:5]}, ensure_ascii=False, indent=2))
