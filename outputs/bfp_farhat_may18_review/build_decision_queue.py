from openpyxl import load_workbook
from pathlib import Path
from collections import defaultdict, Counter
import csv, re, difflib, json

NEW = Path('/Users/tai/Downloads/BFP Main List - May 18 3am Share File Farhat.xlsx')
SYS = Path('/Users/tai/Downloads/guests-2026-05-17-658.csv')
OUT = Path('/Users/tai/Documents/GitHub/taiyzun.com/outputs/bfp_farhat_may18_review')

CONTROLLED = {
    'Artist & Literature','Business','City','Education','Environment','Health','Housing',
    'Influencers','Justice & Peace','Social Innovation','Sports','Youth & Women'
}
AWARD_FIX = {'Champion Business': 'Business', 'Peace': 'Justice & Peace'}
HON = re.compile(r"^(mr|mrs|ms|miss|dr|prof|professor|shri|smt|hon'?ble|hon|he|h\.e\.|h\.h\.|hh|sr\.?|fr\.?|rev\.?|ambassador|maharaja|princess|prince|ca|adv|padma shri|rt rev|most venerable|respected sardar)\.?\s+", re.I)
BUL = re.compile(r"^[\s•⁠\-–—]+")

def clean(x): return '' if x is None else str(x).replace('\u00a0',' ').strip()
def norm_name(x):
    s=clean(x).lower(); s=BUL.sub('',s); s=s.replace('&',' and ')
    s=re.sub(r"\([^)]*\)"," ",s)
    for _ in range(5): s=HON.sub('',s).strip()
    s=re.sub(r"[^a-z0-9]+"," ",s); s=re.sub(r"\s+"," ",s).strip()
    return s

def phone_norm(x):
    s=clean(x)
    digits=re.sub(r'\D+','',s)
    if not digits: return ''
    if len(digits)==10: return '+91'+digits
    return '+'+digits

def load_fans(path):
    wb=load_workbook(path, data_only=True, read_only=True); ws=wb.active
    headers=[c.value if c.value is not None else f'__blank_{i+1}' for i,c in enumerate(next(ws.iter_rows(min_row=1,max_row=1)))]
    rows=[]
    for idx,r in enumerate(ws.iter_rows(min_row=2, values_only=True),2):
        d={headers[i]:clean(r[i]) if i<len(r) else '' for i in range(len(headers))}
        if clean(d.get('Name')):
            d['_row']=idx; d['_key']=norm_name(d.get('Name For Sort') or d.get('Name')); d['_phone_norm']=phone_norm(d.get('Cell'))
            rows.append(d)
    return rows

def load_system(path):
    with open(path, newline='', encoding='utf-8-sig') as f: rows=list(csv.DictReader(f))
    for r in rows:
        r['_key']=norm_name(r.get('name')); r['_phone_norm']=phone_norm(r.get('phone'))
    return rows

def target_from_fans(r):
    cf, cat, award = r.get('Category Final',''), r.get('Category',''), r.get('Award Category (where applicable)','')
    target_status=''; target_champion=''; note=''
    if cf == 'Guest of Honour':
        if cat == 'Awardee' or award:
            target_status='I Am Peacekeeper Champion'
            target_champion=AWARD_FIX.get(award, award)
        else:
            target_status='Guest of Honour'
    elif cf == 'VIP Peace Partner':
        target_status='VIP I Am Peacekeeper Partner'
    elif cf == 'I Am Peacekeeper':
        target_status='KEEP_EXISTING_IAP_STATUS'
        note='Farhat gives broad I Am Peacekeeper only; keep existing COPE/SCOPE/ROPE unless user confirms.'
    elif cf == 'Main':
        if cat == 'Chief Guest': target_status='Chief Guest'
        elif cat == 'Nobel Laureate': target_status='Nobel Peace Laureates'
        else: target_status='KEEP_EXISTING_OR_MANUAL'
    elif cf == 'Media':
        target_status='Media/manual only'
    elif cf == 'XXX':
        target_status='DO_NOT_UPDATE_OR_ADD'
    elif cf == 'Late Peace Guest':
        target_status='Late Peace Guest/manual'
    else:
        target_status='MANUAL'
    return target_status, target_champion, note

fans=load_fans(NEW); system=load_system(SYS)
sys_by_key=defaultdict(list); sys_by_phone=defaultdict(list)
for s in system:
    if s['_key']: sys_by_key[s['_key']].append(s)
    if s['_phone_norm']: sys_by_phone[s['_phone_norm']].append(s)
all_sys_keys=list(sys_by_key)

# duplicate keys in fans
fans_by_key=defaultdict(list)
for f in fans: fans_by_key[f['_key']].append(f)
dup_keys={k for k,v in fans_by_key.items() if len(v)>1}

rows=[]
for f in fans:
    name=f['Name']; key=f['_key']; cf=f.get('Category Final',''); cat=f.get('Category',''); award=f.get('Award Category (where applicable)','')
    target_status,target_champion,note=target_from_fans(f)
    reasons=[]; action=''
    match_type=''; sys_id=''; sys_name=''; sys_status=''; sys_phone=''
    matches=[]
    if key in sys_by_key:
        matches=sys_by_key[key]; match_type='exact_name'
    elif f['_phone_norm'] and f['_phone_norm'] in sys_by_phone:
        matches=sys_by_phone[f['_phone_norm']]; match_type='phone'
    else:
        close=difflib.get_close_matches(key, all_sys_keys, n=3, cutoff=0.88)
        if close:
            matches=[sys_by_key[c][0] for c in close]; match_type='fuzzy_name'
    if len(matches)==1:
        s=matches[0]; sys_id=s.get('id',''); sys_name=s.get('name',''); sys_status=s.get('status_label',''); sys_phone=s.get('phone','')
    elif len(matches)>1:
        reasons.append('multiple system matches')
        sys_id=' | '.join(m.get('id','') for m in matches[:3]); sys_name=' | '.join(m.get('name','') for m in matches[:3]); match_type=match_type+'_multiple'
    else:
        reasons.append('not found in latest system CSV snapshot')

    if cf in ('XXX','Media'):
        action='SKIP_MANUAL_ONLY'; reasons.append(f'{cf} row')
    elif key in dup_keys:
        action='ASK_DUPLICATE'; reasons.append('duplicate normalised name in Farhat list')
    elif match_type.startswith('fuzzy'):
        action='ASK_FUZZY_MATCH'; reasons.append('fuzzy system match')
    elif not sys_id:
        action='ASK_ADD_OR_FIND_LIVE'; reasons.append('no reliable live System ID from CSV snapshot')
    elif target_status in ('KEEP_EXISTING_IAP_STATUS','KEEP_EXISTING_OR_MANUAL','Late Peace Guest/manual','MANUAL'):
        action='ASK_STATUS'; reasons.append(target_status)
    elif cat=='Awardee' and not award:
        action='ASK_AWARD_CATEGORY'; reasons.append('Awardee has no Champion category')
    elif award and award not in CONTROLLED and award not in AWARD_FIX:
        action='ASK_AWARD_WORDING'; reasons.append('award category not controlled')
    else:
        action='READY_SAFE_UPDATE'
    if cf=='I Am Peacekeeper' and clean(f.get('Seat 18th 3am')):
        reasons.append('IAP has seat in Farhat sheet; verbal says no seat')
    if note: reasons.append(note)

    rows.append({
        'Action': action, 'Reasons': '; '.join(dict.fromkeys(reasons)), 'FANS Row': f['_row'], 'FANS Name': name,
        'FANS Category Final': cf, 'FANS Category': cat, 'FANS Award Category': award,
        'Target Status': target_status, 'Target Champion Category': target_champion,
        'FANS Seat 18th 3am': f.get('Seat 18th 3am',''), 'FANS Cell': f.get('Cell',''), 'FANS Email': f.get('email',''),
        'Match Type': match_type, 'System ID': sys_id, 'System Name': sys_name, 'System Status': sys_status, 'System Phone': sys_phone,
    })

headers=list(rows[0].keys())
with open(OUT/'farhat_may18_decision_queue.csv','w',newline='',encoding='utf-8-sig') as f:
    w=csv.DictWriter(f, fieldnames=headers); w.writeheader(); w.writerows(rows)
for action in sorted(set(r['Action'] for r in rows)):
    subset=[r for r in rows if r['Action']==action]
    safe_name='queue_'+re.sub(r'[^A-Za-z0-9]+','_',action).strip('_').lower()+'.csv'
    with open(OUT/safe_name,'w',newline='',encoding='utf-8-sig') as f:
        w=csv.DictWriter(f, fieldnames=headers); w.writeheader(); w.writerows(subset)
summary={'total_rows':len(rows),'actions':Counter(r['Action'] for r in rows),'target_statuses':Counter(r['Target Status'] for r in rows),'ready_safe_update':sum(1 for r in rows if r['Action']=='READY_SAFE_UPDATE')}
(OUT/'decision_queue_summary.json').write_text(json.dumps(summary,default=lambda x:dict(x),indent=2),encoding='utf-8')
print(json.dumps(summary,default=lambda x:dict(x),indent=2))
print('\nFirst manual rows:')
for r in rows:
    if r['Action']!='READY_SAFE_UPDATE':
        print(r['Action'], r['FANS Row'], r['FANS Name'], '|', r['Reasons'], '| target', r['Target Status'], r['Target Champion Category'], '| match', r['Match Type'], r['System Name'])
